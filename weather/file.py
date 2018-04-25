import openpyxl
import datetime
import xlsxwriter

xfile = openpyxl.load_workbook('/home/mugunth/Desktop/allweather.xlsx')


sheet = xfile.get_sheet_by_name('Sheet1')

tmp="A"+str(sheet.max_row)
date=datetime.date.today()
#date = date.strftime("%d/%m/%Y")
#sheet[tmp]=datetime.date.today().strftime("%d/%m/%Y")
#xfile.save('/home/mugunth/Desktop/allweather.xlsx')
'''
if d1==date:
    print("yeah")
    tmp="A"+str(sheet.max_row+1)
    sheet[tmp]=datetime.date.today().strftime("%d/%m/%Y")
    xfile.save('/home/mugunth/Desktop/weather/allweather.xlsx')
'''
